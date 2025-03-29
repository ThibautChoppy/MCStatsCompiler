#TODO: manage correctly importdatas etc based on enabled features

import json
import os
import pandas as pd
import numpy as np
import configparser
import openpyxl
import datetime
import ftplib
import math
import warnings
import paramiko
import sqlite3
import stat
import sqlite3
import shutil
import nbt
from PIL import Image, ImageDraw, ImageFont, UnidentifiedImageError
import requests
from io import BytesIO
import networkx as nx
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import matplotlib.ticker as ticker
import matplotlib.font_manager as fm


# Creation or update of the SQLite table
def init_database(db_path):
    try:
        print(f"Initialisation de la base de données à {db_path}...")
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # Table creation for leaderboards
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS standard_leaderboard (
                rank INTEGER,
                player_name TEXT,
                score INTEGER,
                last_updated TEXT
            )
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS shiny_leaderboard (
                rank INTEGER,
                player_name TEXT,
                score INTEGER,
                last_updated TEXT
            )
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS legendary_leaderboard (
                rank INTEGER,
                player_name TEXT,
                score INTEGER,
                last_updated TEXT
            )
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS money_leaderboard (
                rank INTEGER,
                player_name TEXT,
                score INTEGER,
                last_updated TEXT
            )
        ''')
        conn.commit()
        print("Base de données initialisée avec succès")
        return conn
    except sqlite3.Error as e:
        print(f"Erreur lors de l'initialisation de la base de données : {e}")
        raise

# List contents of directory and parent directory for debugging
def list_sftp_directory(sftp, path="."):
    try:
        print(f"\nContents of current directory '{path}':")
        for entry in sftp.listdir_attr(path):
            print(f"{entry.filename:30} {'<DIR>' if stat.S_ISDIR(entry.st_mode) else '<FILE>'}")
        
        parent = os.path.dirname(path) if path != "/" else "/"
        print(f"\nContents of parent directory '{parent}':")
        for entry in sftp.listdir_attr(parent):
            print(f"{entry.filename:30} {'<DIR>' if stat.S_ISDIR(entry.st_mode) else '<FILE>'}")
    except Exception as e:
        print(f"Error listing directory: {e}")


def loadVanillaData(csvtoggle, csvpath, inputmode, ftpserver, ftppath, localpath, csvtogglemoney, csvpathmoney):
    df = pd.DataFrame()
    money = {}
    waystones = {}
    advancements = pd.DataFrame()
    
    if inputmode == "ftp" or inputmode == "sftp":
        
        if ftppath == "":
            ftppath_complete_stats = "world/stats"
            ftppath_complete_playerdata = "world/playerdata"
            ftppath_complete_advancements = "world/advancements"
        else:
            ftppath_complete_stats = ftppath + "/world/stats"
            ftppath_complete_playerdata = ftppath + "/world/playerdata"
            ftppath_complete_advancements = ftppath + "/world/advancements"
        if inputmode == "ftp":
            ftpserver.cwd(ftppath)
            with open("data/usercache/usercache.json", "wb") as file:
                ftpserver.retrbinary(f"RETR usercache.json", file.write)
            names = pd.DataFrame(json.load(open("data/usercache/usercache.json", "r")))
            # Go back to root
            ftpserver.cwd("../" * (len(ftpserver.pwd().split("/"))-1))
            # Get directories
            filenames = ftpserver.nlst(ftppath_complete_stats)
            ftpserver.cwd(ftppath_complete_stats)
        elif inputmode == "sftp":
            try:
                ftpserver.chdir(ftppath)
            except IOError:
                print(f"Failed to change to directory {ftppath}")
                list_sftp_directory(ftpserver)
                raise
            try:
                ftpserver.get("usercache.json", "data/usercache/usercache.json")
            except IOError:
                print("Failed to get usercache.json")
                list_sftp_directory(ftpserver)
                raise
            names = pd.DataFrame(json.load(open("data/usercache/usercache.json", "r")))
            try:
                current_path = ftpserver.getcwd()
                depth = len([x for x in current_path.split("/") if x]) if current_path != "/" else 0
                if depth > 0:
                    ftpserver.chdir("../" * depth)  # Return to root
                print(f"Trying to access {ftppath_complete_stats}")
                filenames = ftpserver.listdir(ftppath_complete_stats)
                ftpserver.chdir(ftppath_complete_stats)
            except IOError:
                print(f"Failed to access {ftppath_complete_stats}")
                list_sftp_directory(ftpserver)
                raise

        # Start by removing current data files in local
        for folder in ["data/stats", "data/playerdata", "data/advancements"]:
            for filename in os.listdir(folder):
                file_path = os.path.join(folder, filename)
                try:
                    if filename == ".gitignore":
                        continue
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)
                except Exception as e:
                    print('Failed to remove %s. Reason: %s' % (file_path, e))

        # Process the stats data
        for filename in filenames:
            if filename[-1] == ".":
                continue
            filename = filename.split("/")[-1]
            print("Now processing", filename)
            # Download the file to process
            local_file = "data/stats/"+filename
            with open(local_file, "wb") as file:
                if inputmode == "ftp":
                    ftpserver.retrbinary(f"RETR {filename}", file.write)
                elif inputmode == "sftp":
                    ftpserver.get(filename, local_file)
            with open(local_file, "r") as file:
                data = json.load(file)
            
            # Import the JSON to a Pandas DF
            temp_df = pd.json_normalize(data, meta_prefix=True)
            temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
            if temp_name.empty:
                print("No username found for UUID", filename[:-5], " in usercache.json, using UUID for this player instead.")
                temp_name = filename[:-5]
                temp_df = temp_df.transpose().iloc[1:].rename({0: temp_name}, axis=1)
            temp_df = temp_df.transpose().iloc[1:].rename({0: temp_name.iloc[0]}, axis=1)
            # Split the index (stats.blabla.blabla) into 3 indexes (stats, blabla, blabla)
            temp_df.index = temp_df.index.str.split('.', expand=True)
            # If a stat name has a dot in it, remove the part after the dot
            if len(temp_df.index.levshape) > 3:
                temp_df.index = temp_df.index.droplevel(3)
                temp_df = temp_df.groupby(level=[0,1,2]).sum()
            #print(temp_df)
            #temp_df.to_csv('temp.csv')
            if df.empty:
                df = temp_df
            else:
                df = df.join(temp_df, how="outer")
        
        # Go back to previous folder, now use playerdata
        if inputmode == "ftp":
            # Go back to root
            ftpserver.cwd("../" * (len(ftpserver.pwd().split("/"))-1))
            # Get directories
            filenames = ftpserver.nlst(ftppath_complete_playerdata)
            ftpserver.cwd(ftppath_complete_playerdata)
        elif inputmode == "sftp":
            # Go back to root
            current_path = ftpserver.getcwd()
            depth = len([x for x in current_path.split("/") if x]) if current_path != "/" else 0
            if depth > 0:
                ftpserver.chdir("../" * depth)  # Return to root
            # Get directories
            filenames = ftpserver.listdir(ftppath_complete_playerdata)
            ftpserver.chdir(ftppath_complete_playerdata)
        for filename in filenames:
            filename = filename.split("/")[-1]
            if filename[-1] == "." or filename[-4:] == "_old" or filename == "player_roles":
                continue
            print("Now processing", filename)
            # Download the file to process
            local_file = "data/playerdata/"+filename
            with open(local_file, "wb") as file:
                if inputmode == "ftp":
                    ftpserver.retrbinary(f"RETR {filename}", file.write)
                elif inputmode == "sftp":
                    ftpserver.get(filename, local_file)
            temp_name = names.loc[names['uuid'] == filename[:-4]]['name']
            nbtfile = nbt.nbt.NBTFile(local_file,'r')
            money[temp_name.iloc[0]] = math.floor(nbtfile['cardinal_components']['numismatic-overhaul:currency']['Value'].value/10000)
            waystones[temp_name.iloc[0]] = len(nbtfile['BalmData']['WaystonesData']['Waystones'])
        money = pd.DataFrame(money, index=["money"]).transpose()
        waystones = pd.DataFrame(waystones, index=["waystones"]).transpose()
        
        # Go back to previous folder, now use advancements
        if inputmode == "ftp":
            # Go back to root
            ftpserver.cwd("../" * (len(ftpserver.pwd().split("/"))-1))
            # Get directories
            filenames = ftpserver.nlst(ftppath_complete_advancements)
            ftpserver.cwd(ftppath_complete_advancements)
        elif inputmode == "sftp":
            # Go back to root
            current_path = ftpserver.getcwd()
            depth = len([x for x in current_path.split("/") if x]) if current_path != "/" else 0
            if depth > 0:
                ftpserver.chdir("../" * depth)  # Return to root
            # Get directories
            filenames = ftpserver.listdir(ftppath_complete_advancements)
            ftpserver.chdir(ftppath_complete_advancements)
        for filename in filenames:
            filename = filename.split("/")[-1]
            if filename[-1] == ".":
                continue
            print("Now processing", filename)
            # Download the file to process
            local_file = "data/advancements/"+filename
            with open(local_file, "wb") as file:
                if inputmode == "ftp":
                    ftpserver.retrbinary(f"RETR {filename}", file.write)
                elif inputmode == "sftp":
                    ftpserver.get(filename, local_file)
            with open(local_file, "r") as file:
                data = json.load(file)
            temp_df = pd.json_normalize(data, meta_prefix=True)
            temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
            if temp_name.empty:
                print("No username found for UUID", filename[:-5], " in usercache.json, using UUID for this player instead.")
                temp_name = filename[:-5]
                temp_df = temp_df.transpose().iloc[1:].rename({0: temp_name}, axis=1)
            temp_df = temp_df.transpose().iloc[1:].rename({0: temp_name.iloc[0]}, axis=1)
            # Split the index (stats.blabla.blabla) into 3 indexes (stats, blabla, blabla)
            temp_df.index = temp_df.index.str.split('.', expand=True)
            #print(temp_df)
            #temp_df.to_csv('temp.csv')
            if advancements.empty:
                advancements = temp_df
            else:
                advancements = advancements.join(temp_df, how="outer")
        
        # Go back to root
        if inputmode == "ftp":
            ftpserver.cwd("../" * (len(ftpserver.pwd().split("/"))-1))
        elif inputmode == "sftp":
            current_path = ftpserver.getcwd()
            depth = len([x for x in current_path.split("/") if x]) if current_path != "/" else 0
            if depth > 0:
                ftpserver.chdir("../" * depth)
    else:
        if inputmode == "manual":
            names_file = open('data/usercache/usercache.json', 'r')
        elif inputmode == "local":
            names_file = open(localpath+'/usercache.json', 'r')
        names = pd.DataFrame(json.load(names_file))
        if inputmode == "manual":
            playerdata_path = 'data/playerdata'
            stats_path = 'data/stats'
            advancements_path = 'data/advancements'
        if inputmode == "local":
            playerdata_path = localpath+'/world/playerdata'
            advancements_path = localpath+'/world/advancements'
            
        # Stats
        for filename in os.listdir(stats_path):
            if filename == ".gitignore":
                continue
            print("Now processing", filename)
            file = open(stats_path + '/' + filename)
            data = json.load(file)
            # Import the JSON to a Pandas DF
            temp_df = pd.json_normalize(data, meta_prefix=True)
            temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
            if temp_name.empty:
                print("No username found for UUID", filename[:-5], " in usercache.json, using UUID for this player instead.")
                temp_name = filename[:-5]
                temp_df = temp_df.transpose().iloc[1:].rename({0: temp_name}, axis=1)
            temp_df = temp_df.transpose().iloc[1:].rename({0: temp_name.iloc[0]}, axis=1)
            # Split the index (stats.blabla.blabla) into 3 indexes (stats, blabla, blabla)
            temp_df.index = temp_df.index.str.split('.', expand=True)
            # If a stat name has a dot in it, remove the part after the dot
            if len(temp_df.index.levshape) > 3:
                temp_df.index = temp_df.index.droplevel(3)
                temp_df = temp_df.groupby(level=[0,1,2]).sum()
            #print(temp_df)
            #temp_df.to_csv('temp.csv')
            if df.empty:
                df = temp_df
            else:
                df = df.join(temp_df, how="outer")
            
        # Userdata
        for filename in os.listdir(playerdata_path):
            filename = filename.split("/")[-1]
            if filename[-1] == "." or filename[-4:] == "_old" or filename == "player_roles" or filename == ".gitignore":
                continue
            print("Now processing", filename)
            temp_name = names.loc[names['uuid'] == filename[:-4]]['name']
            nbtfile = nbt.nbt.NBTFile(playerdata_path + '/' + filename,'r')
            money[temp_name.iloc[0]] = math.floor(nbtfile['cardinal_components']['numismatic-overhaul:currency']['Value'].value/10000)
            waystones[temp_name.iloc[0]] = len(nbtfile['BalmData']['WaystonesData']['Waystones'])
        money = pd.DataFrame(money, index=["money"]).transpose()
        waystones = pd.DataFrame(waystones, index=["waystones"]).transpose()
            
        # Advancements
        for filename in os.listdir(advancements_path):
            if filename == ".gitignore":
                continue
            print("Now processing", filename)
            file = open(advancements_path + '/' + filename)
            data = json.load(file)
            # Import the JSON to a Pandas DF
            temp_df = pd.json_normalize(data, meta_prefix=True)
            temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
            if temp_name.empty:
                print("No username found for UUID", filename[:-5], " in usercache.json, using UUID for this player instead.")
                temp_name = filename[:-5]
                temp_df = temp_df.transpose().iloc[1:].rename({0: temp_name}, axis=1)
            temp_df = temp_df.transpose().iloc[1:].rename({0: temp_name.iloc[0]}, axis=1)
            # Split the index (stats.blabla.blabla) into 3 indexes (stats, blabla, blabla)
            temp_df.index = temp_df.index.str.split('.', expand=True)
            # Drop rows where it's not a real advancement, it's only an unlocked recipe (i.e. takes the shape of xxx:recipes/yyy)
            temp_df = temp_df[~temp_df.index.get_level_values(0).str.split(":", n=1).str[1].str.startswith("recipes")]
            #print(temp_df)
            #temp_df.to_csv('temp.csv')
            if advancements.empty:
                advancements = temp_df
            else:
                advancements = advancements.join(temp_df, how="outer")
    
    # Replace missing values by 0 (the stat has simply not been initialized because the associated action was not performed)
    df = df.fillna(0)
    advancements = advancements.fillna(0)
    if csvtoggle == "true":
        df.to_csv(csvpath)
    if csvtogglemoney == "true":
        money.to_csv(csvpathmoney)
    return df, money, waystones, advancements

def loadCobblemonData(csvtoggle, csvpath, inputmode, ftpserver, ftppath, localpath):
    # Contains cobbledex_discovery/registers
    df = pd.DataFrame()
    # Contains PvP and PvW victory counts
    df2 = pd.DataFrame()
    # Contains captureCount/defeats
    df3 = pd.DataFrame()
    # Contains PvP duels
    df4 = pd.DataFrame()
    # Contains totalTypeCaptureCounts
    df5 = pd.DataFrame()
    root_dirnames = []
    
    if inputmode == "ftp" or inputmode == "sftp":
        if ftppath == "":
            ftppath_complete = "world/cobblemonplayerdata"
        else:
            ftppath_complete = ftppath + "/world/cobblemonplayerdata"
        if inputmode == "ftp":
            ftpserver.cwd(ftppath)
            with open("data/usercache/usercache.json", "wb") as file:
                ftpserver.retrbinary(f"RETR usercache.json", file.write)
            names = pd.DataFrame(json.load(open("data/usercache/usercache.json", "r")))
            # Go back to root
            ftpserver.cwd("../" * (len(ftpserver.pwd().split("/"))-1))
            # Get directories
            root_dirnames = ftpserver.nlst(ftppath_complete)
            ftpserver.cwd(ftppath_complete)
        else:
            try:
                ftpserver.chdir(ftppath)
            except IOError:
                print(f"Failed to change to directory {ftppath}")
                list_sftp_directory(ftpserver)
                raise
            try:
                ftpserver.get("usercache.json", "data/usercache/usercache.json")
            except IOError:
                print("Failed to get usercache.json")
                list_sftp_directory(ftpserver)
                raise
            names = pd.DataFrame(json.load(open("data/usercache/usercache.json", "r")))
            try:
                current_path = ftpserver.getcwd()
                depth = len([x for x in current_path.split("/") if x]) if current_path != "/" else 0
                if depth > 0:
                    ftpserver.chdir("../" * depth)  # Return to root
                print(f"Trying to access {ftppath_complete}")
                root_dirnames = ftpserver.listdir(ftppath_complete)
                ftpserver.chdir(ftppath_complete)
            except IOError:
                print(f"Failed to access {ftppath_complete}")
                list_sftp_directory(ftpserver)
                raise
        
        # Start by removing current data files in local
        for filename in os.listdir("data/cobblemonplayerdata"):
            file_path = os.path.join("data/cobblemonplayerdata", filename)
            try:
                if filename == ".gitignore":
                    continue
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                print('Failed to remove %s. Reason: %s' % (file_path, e))
        for dirname in root_dirnames:
            if dirname[-1] == ".":
                continue
            subfolder = dirname.split("/")[-1]
            # Go to the subfolder
            if inputmode == "ftp":
                ftpserver.cwd(dirname.split("/")[-1])
                filenames = ftpserver.nlst()
            else:
                ftpserver.chdir(dirname.split("/")[-1])
                filenames = ftpserver.listdir()
            
            # Create the sub-folder on the local level
            os.mkdir("data/cobblemonplayerdata/"+subfolder)
            for filename in filenames:
                if filename == "." or filename == "..":
                    continue
                print("Now processing", filename)
                
                # Download the file to process
                local_file = "data/cobblemonplayerdata/"+subfolder+"/"+filename
                with open(local_file, "wb") as file:
                    if inputmode == "ftp":
                        ftpserver.retrbinary(f"RETR {filename}", file.write)
                    else:
                        ftpserver.get(filename, local_file)
                
                with open(local_file, "r") as file:
                    json_file = json.load(file)
                    data = json_file['extraData']['cobbledex_discovery']['registers']
                    advancementData = json_file['advancementData']
                    captureCountData = json_file['extraData']['captureCount']['defeats']
                    try:
                        duelsData = json_file['extraData']['cobblenavContactData']['contacts']
                    except KeyError:
                        duelsData = None
                
                temp_df = pd.json_normalize(data, meta_prefix=True)
                temp_df = temp_df.transpose().iloc[:]
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                if temp_name.empty:
                    print("No username found for UUID", filename[:-5], " in usercache.json, using UUID for this player instead.")
                    temp_name = filename[:-5]
                    temp_df = temp_df.rename({0: temp_name}, axis=1)
                else:
                    temp_df = temp_df.rename({0: temp_name.iloc[0]}, axis=1)
                if not temp_df.empty:
                    temp_df.index = temp_df.index.str.split('.', expand=True)
                    if df.empty:
                        df = temp_df
                    else:
                        df = df.join(temp_df, how="outer")
                else:
                    df[temp_name] = np.nan
                    
                df2.loc["totalPvPBattleVictoryCount", temp_name] = advancementData['totalPvPBattleVictoryCount']
                df2.loc["totalPvWBattleVictoryCount", temp_name] = advancementData['totalPvWBattleVictoryCount']
                
                temp_df = pd.json_normalize(captureCountData, meta_prefix=True)
                temp_df = temp_df.transpose().iloc[:]
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                if temp_name.empty:
                    temp_name = filename[:-5]
                    temp_df = temp_df.rename({0: temp_name}, axis=1)
                else:
                    temp_df = temp_df.rename({0: temp_name.iloc[0]}, axis=1)
                if not temp_df.empty:
                    if df3.empty:
                        df3 = temp_df
                    else:
                        df3 = df3.join(temp_df, how="outer")
                else:
                    df3[temp_name] = np.nan
                
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                if duelsData != None:
                    temp_df = pd.json_normalize(duelsData, meta_prefix=True)
                    temp_df = temp_df.transpose().iloc[:]
                    temp_df = pd.DataFrame(temp_df.T.stack())
                else:
                    temp_df = pd.DataFrame()
                if temp_name.empty:
                    temp_name = filename[:-5]
                    temp_df = temp_df.rename({0: temp_name}, axis=1)
                else:
                    temp_df = temp_df.rename({0: temp_name.iloc[0]}, axis=1)
                if not temp_df.empty:
                    if df4.empty:
                        df4 = temp_df
                    else:
                        df4 = df4.join(temp_df, how="outer")
                else:
                    df4[temp_name] = np.nan
                
                temp_df = pd.json_normalize(advancementData['totalTypeCaptureCounts'], meta_prefix=True)
                temp_df = temp_df.transpose().iloc[:]
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                if temp_name.empty:
                    temp_name = filename[:-5]
                    temp_df = temp_df.rename({0: temp_name}, axis=1)
                else:
                    temp_df = temp_df.rename({0: temp_name.iloc[0]}, axis=1)
                if not temp_df.empty:
                    if df5.empty:
                        df5 = temp_df
                    else:
                        df5 = df5.join(temp_df, how="outer")
                else:
                    df5[temp_name] = np.nan
                
            if inputmode == "ftp":
                ftpserver.cwd("../")  # Move back to the parent directory
            else:
                ftpserver.chdir("..")
        # Go back to root
        if inputmode == "ftp":
            ftpserver.cwd("../" * (len(ftpserver.pwd().split("/"))-1))
        else:
            current_path = ftpserver.getcwd()
            depth = len([x for x in current_path.split("/") if x]) if current_path != "/" else 0
            if depth > 0:
                ftpserver.chdir("../" * depth)
    else:
        if inputmode == "manual":
            names_file = open('data/usercache/usercache.json', 'r')
        elif inputmode == "local":
            names_file = open(localpath+'/usercache.json', 'r')
        names = pd.DataFrame(json.load(names_file))
        if inputmode == "manual":
            path = 'data/cobblemonplayerdata'
        if inputmode == "local":
            path = localpath+'/world/cobblemonplayerdata'
        i = -1
        for dirpath, dirnames, filenames in os.walk(path):
            if len(dirnames) > 0:
                root_dirnames = dirnames
            for filename in filenames:
                if filename == ".gitignore":
                    continue
                print("Now processing", filename)
                file = open(path + '/' + root_dirnames[i] + '/' + filename)
                json_file = json.load(file)
                data = json_file['extraData']['cobbledex_discovery']['registers']
                advancementData = json_file['advancementData']
                captureCountData = json_file['extraData']['captureCount']['defeats']
                try:
                    duelsData = json_file['extraData']['cobblenavContactData']['contacts']
                except KeyError:
                    duelsData = None
                    
                # Import the JSON to a Pandas DF
                temp_df = pd.json_normalize(data, meta_prefix=True)
                temp_df = temp_df.transpose().iloc[:]
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                if temp_name.empty:
                    print("No username found for UUID", filename[:-5], " in usercache.json, using UUID for this player instead.")
                    temp_name = filename[:-5]
                    temp_df = temp_df.rename({0: temp_name}, axis=1)
                else:
                    temp_df = temp_df.rename({0: temp_name.iloc[0]}, axis=1)
                if not temp_df.empty:
                    # Split the index (stats.blabla.blabla) into 3 indexes (stats, blabla, blabla)
                    temp_df.index = temp_df.index.str.split('.', expand=True)
                    if df.empty:
                        df = temp_df
                    else:
                        df = df.join(temp_df, how="outer")
                else:
                    df[temp_name] = np.nan
                    
                df2.loc["totalPvPBattleVictoryCount", temp_name] = advancementData['totalPvPBattleVictoryCount']
                df2.loc["totalPvWBattleVictoryCount", temp_name] = advancementData['totalPvWBattleVictoryCount']
                df2.loc["totalTradedCount", temp_name] = advancementData['totalTradedCount']
                
                temp_df = pd.json_normalize(captureCountData, meta_prefix=True)
                temp_df = temp_df.transpose().iloc[:]
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                if temp_name.empty:
                    temp_name = filename[:-5]
                    temp_df = temp_df.rename({0: temp_name}, axis=1)
                else:
                    temp_df = temp_df.rename({0: temp_name.iloc[0]}, axis=1)
                if not temp_df.empty:
                    if df3.empty:
                        df3 = temp_df
                    else:
                        df3 = df3.join(temp_df, how="outer")
                else:
                    df3[temp_name] = np.nan
                
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                if duelsData != None:
                    temp_df = pd.json_normalize(duelsData, meta_prefix=True)
                    temp_df = temp_df.transpose().iloc[:]
                    temp_df = pd.DataFrame(temp_df.T.stack())
                else:
                    temp_df = pd.DataFrame()
                if temp_name.empty:
                    temp_name = filename[:-5]
                    temp_df = temp_df.rename({0: temp_name}, axis=1)
                else:
                    temp_df = temp_df.rename({0: temp_name.iloc[0]}, axis=1)
                if not temp_df.empty:
                    if df4.empty:
                        df4 = temp_df
                    else:
                        df4 = df4.join(temp_df, how="outer")
                else:
                    df4[temp_name] = np.nan
                    
                temp_df = pd.json_normalize(advancementData['totalTypeCaptureCounts'], meta_prefix=True)
                temp_df = temp_df.transpose().iloc[:]
                temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
                if temp_name.empty:
                    temp_name = filename[:-5]
                    temp_df = temp_df.rename({0: temp_name}, axis=1)
                else:
                    temp_df = temp_df.rename({0: temp_name.iloc[0]}, axis=1)
                if not temp_df.empty:
                    if df5.empty:
                        df5 = temp_df
                    else:
                        df5 = df5.join(temp_df, how="outer")
                else:
                    df5[temp_name] = np.nan
                
            i += 1
    # Replace missing values by 0 (the stat has simply not been initialized because the associated action was not performed)
    df = df.fillna(0)
    df3 = df3.fillna(0)
    df4 = df4.fillna(0)
    df5 = df5.fillna(0)
    if csvtoggle == "true":
        df.to_csv(csvpath)
    return df, df2, df3, df4, df5


def getVanillaLeaderboard(df, cat, subcat, verbose=True):
    if subcat == "total":
        row = df.loc['stats'].loc[cat].sum().sort_values().iloc[::-1]
    else:
        row = df.loc['stats'].loc[cat].loc[subcat].sort_values().iloc[::-1]
    df = pd.DataFrame(row).rename(columns={subcat: 0})
    if cat == "minecraft:custom" and subcat == "minecraft:play_time":
        df[0] = df[0].apply(lambda x: f"{(int(x) // (20*60*60))}h {((int(x)) // (20*60))%60}min")
    if verbose:
        print("Leaderboard of", cat, subcat, ":")
        print(df)
    return df

def getVanillaBestAndWorst(df, username, cleaning, cleaningvalue):
    if username == "null" or not username:
        print("Error for Best-and-Worst feature: no username specified in the config")
        return
        
    if username not in df.columns:
        print(f"Error for Best-and-Worst feature: User '{username}' does not exist in the provided data")
        print("Available users:", ", ".join(df.columns))
        return
        
    nb_players = df.shape[1]
    if cleaning == "true":
        before_value = df.shape[0]
        df['zero_count'] = df.apply(lambda row: (row == 0).sum(), axis=1)
        df = df.drop(df[df['zero_count'] > (nb_players-int(cleaningvalue))].index)
        df = df.drop('zero_count', axis=1)
        print(before_value - df.shape[0], "rows dropped out of", before_value, "because of cleaning.")
    ranks = df.rank(axis=1, method='min', ascending=False)
    ranks['non_zero_values'] = df.apply(lambda row: nb_players - (row == 0).sum(), axis=1)
    ranks['value'] = df[username]
    output = ranks[[username, 'value', 'non_zero_values']].sort_values(username, ascending=False).rename(columns={username:"rank_"+username, "value":"value_"+username})
    with pd.option_context('display.max_rows', None):
        print(output) # add .to_string() for the whole output

def getAdvancementsLeaderboard(df):
    count_df = pd.DataFrame((df == True).sum().sort_values())
    count_df['index'] = range(len(count_df), 0, -1)
    count_df = count_df.iloc[::-1]
    return count_df

def getStandardLeaderboard(df):
    row = df.sort_values().iloc[::-1]
    df = pd.DataFrame(row).rename(columns={type: 0})
    return df

def getCobblemonCaptureCountLeaderboard(df):
    df = df.reset_index().melt(id_vars='index')
    df.columns = ['cobblemon', 'player', 'value']
    df = df.sort_values(by='value', ascending=False).set_index('player')
    return df
    
def most_pokemons_leaderboard(df, config, type, conn):
    if config['COBBLEMONLEADERBOARDS']['SQLiteOutput'] == "true":
        try:
            print(f"\nInsertion des données dans la table {type}...")
            print(f"Données à insérer :\n{df.head()}")
            print(f"Nombre de lignes : {len(df)}")
            cursor = conn.cursor()
            # Table selection by leaderboard type
            table_map = {
                "standard": "standard_leaderboard",
                "shiny": "shiny_leaderboard",
                "legendary": "legendary_leaderboard",
                "money": "money_leaderboard"
            }
            table_name = table_map[type]
            print(f"Table cible : {table_name}")
            # Clear old data
            cursor.execute(f"DELETE FROM {table_name}")
            # New data insertion
            now = datetime.datetime.now().strftime(config['COBBLEMONLEADERBOARDS']['LastUpdated'])
            print(f"Date de mise à jour : {now}")
            for idx, row in df.iterrows():
                try:
                    rank = int(row['index']) if 'index' in row else int(row[0])
                    score = int(row.iloc[0]) if 'index' in row else int(row[1])
                    cursor.execute(f'''
                        INSERT INTO {table_name} (rank, player_name, score, last_updated)
                        VALUES (?, ?, ?, ?)
                    ''', (rank, idx, score, now))
                except sqlite3.Error as e:
                    print(f"Erreur lors de l'insertion de la ligne {idx}: {e}")
                    continue
            conn.commit()
            print(f"Données insérées avec succès dans la table {type}")
            
            # Vérification des données insérées
            cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
            count = cursor.fetchone()[0]
            print(f"Nombre de lignes dans la table {table_name}: {count}")
            cursor.execute(f"SELECT * FROM {table_name} LIMIT 5")
            print(f"Premières lignes de la table {table_name}:")
            for row in cursor.fetchall():
                print(row)
        except sqlite3.Error as e:
            print(f"Erreur lors de l'opération sur la base de données : {e}")
            conn.rollback()
            raise
    
    if config['COBBLEMONLEADERBOARDS']['XLSXOutput'] == "true":
        # Load the Excel file
        file_path = "output.xlsx"
        wb = openpyxl.load_workbook(file_path)
        tab_map = {
            "standard": "leaderboard2",
            "shiny": "leaderboard3",
            "legendary": "leaderboard4",
            "money": "leaderboard5"
        }
        ws = wb[tab_map[type]]
        i = 0
        ExcelRows = int(config['COBBLEMONLEADERBOARDS']['ExcelRows'])
        ExcelCols = int(config['COBBLEMONLEADERBOARDS']['ExcelColumns'])
        for index, row in df[0:ExcelRows*ExcelCols].iterrows():
            ws.cell(row=(i%ExcelRows)+3, column=2+math.floor(i/ExcelRows)*3, value=str(i+1)+".")
            ws.cell(row=(i%ExcelRows)+3, column=3+math.floor(i/ExcelRows)*3, value=index)
            ws.cell(row=(i%ExcelRows)+3, column=4+math.floor(i/ExcelRows)*3, value=row.iloc[0])
            i += 1
        now = datetime.datetime.now()
        ws.cell(row=ExcelRows+3, column=2, value=now.strftime(config['COBBLEMONLEADERBOARDS']['LastUpdated']))
        ws.cell(row=ExcelRows+4, column=2, value=config['COBBLEMONLEADERBOARDS']['Subtitle'])
        wb.save(file_path)

def top_image(df_list, config, titles, special_list):
    translations_df = pd.read_csv('staticdata/pokemon_translations.csv')
    ldb_width = int(config["TOPIMAGE"]["Width"])
    ldb_height = int(config["TOPIMAGE"]["Height"])
    base_y_offset = 50
    base_width, base_height = 640, 210+int(config['TOPIMAGE']['NbPlayers'])*base_y_offset
    width, height = base_width*ldb_width, base_height*ldb_height
    # Load background image
    background = Image.open("images/background4.png").convert("RGBA")
    background = background.resize((width, height), Image.Resampling.LANCZOS)
    img = background.copy()
    
    # Import usernames to show instead of the Minecraft username
    leaderboard_usernames_df = pd.read_csv('staticdata/leaderboard_usernames.csv')
    
    for i, df in enumerate(df_list):
        x_margin = i%ldb_width * base_width
        y_margin = math.floor(i/ldb_height) * base_height
        # Add the rounded rectangles
        overlay = Image.new('RGBA', img.size)
        draw2 = ImageDraw.Draw(overlay)
        draw2.rounded_rectangle([x_margin+10, y_margin+5, x_margin+base_width-10, y_margin+base_height-5], radius=40, fill=(0, 0, 0, 100))
        img = Image.alpha_composite(img, overlay)
        # Title
        font = ImageFont.truetype("fonts/Minecraft-Seven_v2.ttf", 32)
        draw = ImageDraw.Draw(img)
        _, _, w, _ = draw.textbbox((0, 0), titles[i], font=font)
        draw.text((((width/ldb_width)-w)/2 + x_margin, 10 + y_margin), titles[i], align="center", fill="gold", font=font)
        
        y_offset = base_y_offset
        rank = 1
        nbplayers = int(config['TOPIMAGE']['NbPlayers'])
        for _, player in df.iloc[:nbplayers].iterrows():
            response = requests.get("https://mc-heads.net/avatar/"+player.name)
            avatar = Image.open(BytesIO(response.content)).resize((64, 64), Image.Resampling.LANCZOS)
            img.paste(avatar, (100 + x_margin, y_offset + y_margin))
            font = ImageFont.truetype("fonts/minecraft.ttf", 24)
            username = leaderboard_usernames_df.loc[leaderboard_usernames_df['minecraft'] == player.name]
            if username.empty:
                username = player.name
            else:
                username = username['real'].iloc[0]
            if special_list[i] == "singletype":
                score = str(int(player.loc['value'])) + ' (' + translations_df.loc[translations_df['en'].apply(str.lower) == player.loc['cobblemon'].lower()]['fr'].iloc[0] + ')'
                response = requests.get("https://cobblemon.tools/pokedex/pokemon/"+player.loc['cobblemon']+'/sprite.png')
                cobblemon_icon = Image.open(BytesIO(response.content)).resize((64, 64), Image.Resampling.NEAREST)
                _, _, w, _ = draw.textbbox((0, 0), f"#{rank} {username} - {score}", font=font)
                img.paste(cobblemon_icon, (180 + w + x_margin, y_offset + y_margin), cobblemon_icon)
            else:
                score = player.iloc[0]
                if not isinstance(score, str):
                    score = int(score)
            draw.text((170 + x_margin, y_offset + 30 + y_margin), f"#{rank} {username} - {score}", fill="white", font=font)
            y_offset += 80
            rank += 1
            
    # Save the final visualization
    img.save(config['TOPIMAGE']['ImagePath'])

def PvP_network(df, config):
    print("Now preparing the duels graph...")
    leaderboard_usernames_df = pd.read_csv('staticdata/leaderboard_usernames.csv')
    mc_font = fm.FontProperties(fname='fonts/Minecraft-Seven_v2.ttf')
    
    # Prepare data
    duels = []
    done = []
    for player in df.columns:
        for encounter in df.index.get_level_values(0).unique():
            encounter_df = df[player].loc[encounter]
            if encounter_df.loc['key'] == 0 or encounter_df.loc['name'] in done:
                continue
            duels.append((player, encounter_df.loc['name'], encounter_df.loc['winnings']+encounter_df.loc['losses']))
        done.append(player)

    G = nx.Graph()
    for p1, p2, count in duels:
        if p1 != p2:  # Avoid self-loops
            G.add_edge(p1, p2, weight=count)
    # Node sizes based on total duel involvement
    node_duel_counts = {player: sum(d['weight'] for _, _, d in G.edges(player, data=True)) for player in G.nodes()}
    node_sizes = np.array([node_duel_counts[player] for player in G.nodes()])
    # Edge widths based on duel frequency
    edge_weights = np.array([d['weight'] for _, _, d in G.edges(data=True)])
    cmap = plt.cm.cool
    norm = mcolors.Normalize(vmin=min(node_sizes), vmax=max(node_sizes))
    node_colors = [cmap(norm(node_duel_counts[player])) for player in G.nodes()]
    pos = nx.forceatlas2_layout(G, strong_gravity=True)
    
    x_vals, y_vals = zip(*pos.values())
    x_pad = (max(x_vals) - min(x_vals)) * 0.1
    y_pad = (max(y_vals) - min(y_vals)) * 0.1
    xlim = (min(x_vals) - x_pad, max(x_vals) + x_pad)
    ylim = (min(y_vals) - y_pad, max(y_vals) + y_pad)
    _, ax = plt.subplots(figsize=(12, 8), facecolor='#000000')  
    ax.set_xlim(xlim)
    ax.set_ylim(ylim)
    ax.imshow(plt.imread('images/background1.png'), extent=(*xlim, *ylim), aspect='auto', zorder=0)
    # Draw edges with varying width and alpha
    nx.draw_networkx_edges(G, pos, width=edge_weights / max(edge_weights) * 5, alpha=0.6, edge_color="gray")
    # Draw nodes with color mapping
    nx.draw_networkx_nodes(G, pos, node_size=(node_sizes / max(node_sizes)) * 2500 + 200, node_color=node_colors, edgecolors="#d1d1d1")
    labels = {player: leaderboard_usernames_df.loc[leaderboard_usernames_df['minecraft'] == player]['real'].iloc[0] if not leaderboard_usernames_df.loc[leaderboard_usernames_df['minecraft'] == player].empty else player for player in G.nodes()}
    #nx.draw_networkx_labels(G, pos, labels, font_size=9, font_color="black", font_weight="bold")
    #nx.draw_networkx_labels(G, pos, labels, font_size=8, font_color="white", font_weight="bold")
    for node, (x, y) in pos.items():
        # Shadow (black, slightly offset)
        ax.text(x + 0.03, y - 0.03, labels[node], fontsize=8, color='black', zorder=2, ha="center", va="center")
        # Main label (white, on top)
        ax.text(x, y, labels[node], fontsize=8, color='white', zorder=3, ha="center", va="center")

    # Color bar
    sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
    sm.set_array([])
    cbar = plt.colorbar(sm, ax=ax, fraction=0.03, pad=0.02)
    cbar.set_label("Duels joués", color="white", font=mc_font)
    cbar.ax.yaxis.set_tick_params(color="white")
    plt.setp(plt.getp(cbar.ax.axes, "yticklabels"), color="white")

    plt.axis("off")
    mc_font_title = fm.FontProperties(fname='fonts/Minecraft-Seven_v2.ttf', size=42)
    plt.title("Réseau des duels", color="white", fontweight="bold", font=mc_font_title)
    plt.savefig(config['PVPNETWORK']['ImagePath'], dpi=299)
    plt.clf()

def cobblemon_types_barchart(df, config):
    print("Now preparing the types barchart...")
    pokemon_types_df = pd.read_csv('staticdata/pokemon_types.csv')
    total = df.sum().sum()
    df['proportion'] = df.apply(lambda row: 100 * row.sum() / total, axis=1)
    df = pd.merge(df, pokemon_types_df, left_on=df.index, right_on='en')
    barchart = plt.bar(df['fr'], df['proportion'], color=df['color'])
    mc_font = fm.FontProperties(fname='fonts/Minecraft-Seven_v2.ttf')
    plt.bar_label(barchart, df['fr'], padding=6, rotation=90, fontsize=18, fontweight='bold', font=mc_font, color='white')
    plt.xticks([], [])
    plt.gca().tick_params(axis='y', colors='white')
    plt.gca().yaxis.set_major_formatter(ticker.PercentFormatter(decimals=0))
    plt.gca().set_facecolor('none') # Set transparent background
    for spine in plt.gca().spines.values(): # Remove the figure border
        spine.set_visible(False)
    plt.savefig(config['TYPESBARCHART']['ImagePath'], dpi=299, transparent=True)
    plt.clf()
    
def stats_pokeballs(config, ftpserver):
    print("Now preparing the pokeball stats...")
    df = pd.DataFrame()
    root_dirnames = []
    
    if config['INPUT']['Mode'] == "ftp" or config['INPUT']['Mode'] == "sftp":
        if config['INPUT']['FTPPath'] == "":
            ftppath_complete = "world/pokemon/pcstore"
        else:
            ftppath_complete = config['INPUT']['FTPPath'] + "/world/pokemon/pcstore"
        if config['INPUT']['Mode'] == "ftp":
            ftpserver.cwd(config['INPUT']['FTPPath'])
            with open("data/usercache/usercache.json", "wb") as file:
                ftpserver.retrbinary(f"RETR usercache.json", file.write)
            names = pd.DataFrame(json.load(open("data/usercache/usercache.json", "r")))
            # Go back to root
            ftpserver.cwd("../" * (len(ftpserver.pwd().split("/"))-1))
            # Get directories
            root_dirnames = ftpserver.nlst(ftppath_complete)
            ftpserver.cwd(ftppath_complete)
        else:
            try:
                ftpserver.chdir(config['INPUT']['FTPPath'])
            except IOError:
                print(f"Failed to change to directory {config['INPUT']['FTPPath']}")
                list_sftp_directory(ftpserver)
                raise
            try:
                ftpserver.get("usercache.json", "data/usercache/usercache.json")
            except IOError:
                print("Failed to get usercache.json")
                list_sftp_directory(ftpserver)
                raise
            names = pd.DataFrame(json.load(open("data/usercache/usercache.json", "r")))
            try:
                current_path = ftpserver.getcwd()
                depth = len([x for x in current_path.split("/") if x]) if current_path != "/" else 0
                if depth > 0:
                    ftpserver.chdir("../" * depth)  # Return to root
                print(f"Trying to access {ftppath_complete}")
                root_dirnames = ftpserver.listdir(ftppath_complete)
                ftpserver.chdir(ftppath_complete)
            except IOError:
                print(f"Failed to access {ftppath_complete}")
                list_sftp_directory(ftpserver)
                raise
        
        # Start by removing current data files in local
        for filename in os.listdir("data/pokemon/pcstore"):
            file_path = os.path.join("data/pokemon/pcstore", filename)
            try:
                if filename == ".gitignore":
                    continue
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                print('Failed to remove %s. Reason: %s' % (file_path, e))
        for filename in os.listdir("data/pokemon/playerpartystore"):
            file_path = os.path.join("data/pokemon/playerpartystore", filename)
            try:
                if filename == ".gitignore":
                    continue
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                print('Failed to remove %s. Reason: %s' % (file_path, e))
        for dirname in root_dirnames:
            if dirname[-1] == ".":
                continue
            subfolder = dirname.split("/")[-1]
            # Go to the subfolder
            if config['INPUT']['Mode'] == "ftp":
                ftpserver.cwd(dirname.split("/")[-1])
                filenames = ftpserver.nlst()
            else:
                ftpserver.chdir(dirname.split("/")[-1])
                filenames = ftpserver.listdir()
            
            # Create the sub-folder on the local level
            os.mkdir("data/pokemon/pcstore/"+subfolder)
            for filename in filenames:
                if filename == "." or filename == "..":
                    continue
                print("Now processing", filename)
                
                # Download the file to process
                local_file = "data/pokemon/pcstore/"+subfolder+"/"+filename
                with open(local_file, "wb") as file:
                    if config['INPUT']['Mode'] == "ftp":
                        ftpserver.retrbinary(f"RETR {filename}", file.write)
                    else:
                        ftpserver.get(filename, local_file)
                
                temp_name = names.loc[names['uuid'] == filename[:-4]]['name']
                nbtfile = nbt.nbt.NBTFile(local_file,'r')
                
                box_count = int(nbtfile['BoxCount'].value)
                balls = {}
                for box in range(box_count):
                    for slot in nbtfile['Box'+str(box)]:
                        ball = nbtfile['Box'+str(box)][slot]['CaughtBall'].value
                        try:
                            balls[ball] += 1
                        except KeyError:
                            balls[ball] = 1
                df[temp_name.iloc[0]] = balls
                
            if config['INPUT']['Mode'] == "ftp":
                ftpserver.cwd("../")  # Move back to the parent directory
            else:
                ftpserver.chdir("..")
        # Go back to root
        if config['INPUT']['Mode'] == "ftp":
            ftpserver.cwd("../" * (len(ftpserver.pwd().split("/"))-1))
        else:
            current_path = ftpserver.getcwd()
            depth = len([x for x in current_path.split("/") if x]) if current_path != "/" else 0
            if depth > 0:
                ftpserver.chdir("../" * depth)
    else:
        if config['INPUT']['Mode'] == "manual":
            names_file = open('data/usercache/usercache.json', 'r')
        elif config['INPUT']['Mode'] == "local":
            names_file = open(config['INPUT']['LocalPath']+'/usercache.json', 'r')
        names = pd.DataFrame(json.load(names_file))
        if config['INPUT']['Mode'] == "manual":
            path = 'data/pokemon/pcstore'
        if config['INPUT']['Mode'] == "local":
            path = config['INPUT']['LocalPath']+'/world/pokemon/pcstore'
        i = -1
        for dirpath, dirnames, filenames in os.walk(path):
            if len(dirnames) > 0:
                root_dirnames = dirnames
            for filename in filenames:
                if filename == ".gitignore":
                    continue
                print("Now processing", filename)
                    
                temp_name = names.loc[names['uuid'] == filename[:-4]]['name']
                nbtfile = nbt.nbt.NBTFile(path + '/' + root_dirnames[i] + '/' + filename,'r')
                
                box_count = int(nbtfile['BoxCount'].value)
                balls = {}
                for box in range(box_count):
                    for slot in nbtfile['Box'+str(box)]:
                        ball = nbtfile['Box'+str(box)][slot]['CaughtBall'].value
                        try:
                            balls[ball] += 1
                        except KeyError:
                            balls[ball] = 1
                df[temp_name.iloc[0]] = balls
                
            i += 1
    # Replace missing values by 0 (the stat has simply not been initialized because the associated action was not performed)
    df = df.fillna(0)
    df.to_csv(config['STATSPOKEBALLS']['CSVPath'])

def player_card(config, username, stats_values):
    cards_df = pd.read_csv('staticdata/cards.csv')
    custom_sentence1 = str(cards_df.loc[cards_df['name'] == username]['stat1'].iloc[0])
    custom_sentence2 = str(cards_df.loc[cards_df['name'] == username]['stat2'].iloc[0])
    starter_fr = str(cards_df.loc[cards_df['name'] == username]['starter_fr'].iloc[0])
    background_path = "images/background5.png"
    skins_path = "images/skins/"
    stats_labels = ["Temps de jeu", "Cobblemons attrapés", "Crates lootées", "Duels PvP gagnés", "Advancements"]

    # Skin image
    skin_img = Image.open(skins_path+username+".png").convert("RGBA")
    skin_img = skin_img.resize((185, 185))

    # Generate radar chart
    size = (400, 300)
    dpi = 299
    num_vars = len(stats_labels)
    angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()
    stats_values += stats_values[:1]
    angles += angles[:1]
    # Calculate figure size in inches for desired pixel size
    fig_width, fig_height = size[0] / dpi, size[1] / dpi
    fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=dpi, subplot_kw=dict(polar=True))
    ax.plot(angles, stats_values, color='cyan', linewidth=2)
    ax.fill(angles, stats_values, color='cyan', alpha=0.25)
    ax.set_ylim(0, 1)
    ax.set_yticklabels([])
    ax.set_xticks(angles[:-1])
    ax.grid(True, linestyle='dotted', color='gray')
    ax.set_xticklabels(stats_labels, color='white', font=fm.FontProperties(fname='fonts/Minecraft-Seven_v2.ttf', size=5))

    # Transparency settings
    fig.patch.set_alpha(0)
    ax.set_facecolor('none')

    radar_buf = BytesIO()
    plt.savefig(radar_buf, format='PNG', transparent=True, bbox_inches='tight', pad_inches=0.1)
    radar_buf.seek(0)
    plt.close()
    radar_chart_img = Image.open(radar_buf)
    radar_chart_img = radar_chart_img.resize(size, Image.Resampling.LANCZOS)

    # Create the card
    card_width, card_height = 700, 450
    if not os.path.exists(background_path):
        raise FileNotFoundError(f"Background image not found at '{background_path}'")
    background_img = Image.open(background_path).convert("RGBA")
    background_img = background_img.resize((card_width, card_height))
    card = background_img.copy()
    draw = ImageDraw.Draw(card)
    # Skin and radar chart images
    card.paste(skin_img, (30, 30), skin_img)
    card.paste(radar_chart_img, (290, 10), radar_chart_img)
    font_title = ImageFont.truetype("fonts/Minecraft-Seven_v2.ttf", 44)
    font_sentence = ImageFont.truetype("fonts/Minecraft-Seven_v2.ttf", 28)

    # Draw username and sentence with a shadow for better visibility
    def draw_text_with_shadow(draw_obj, position, text, font, fill, shadow_color=(0, 0, 0)):
        x, y = position
        draw_obj.text((x+1, y+1), text, font=font, fill=shadow_color)
        draw_obj.text((x, y), text, font=font, fill=fill)
    leaderboard_usernames_df = pd.read_csv('staticdata/leaderboard_usernames.csv')
    display_name = leaderboard_usernames_df.loc[leaderboard_usernames_df['minecraft'] == username]
    if not display_name.empty:
        display_name = display_name['real'].iloc[0]
    else:
        display_name = username
    draw_text_with_shadow(draw, (30, 250), display_name, font_title, fill=(255, 255, 255))
    draw_text_with_shadow(draw, (30, 310), custom_sentence1, font_sentence, fill=(200, 200, 200))
    draw_text_with_shadow(draw, (30, 350), custom_sentence2, font_sentence, fill=(200, 200, 200))
    draw_text_with_shadow(draw, (30, 400), "Starter : "+starter_fr, font_sentence, fill=(230, 230, 230))
    translations_df = pd.read_csv('staticdata/pokemon_translations.csv')
    if starter_fr != "nan":
        starter_en = translations_df.loc[translations_df['fr'] == starter_fr]['en'].iloc[0]
        # Tailow is spelled Taillow on the Cobblemon website
        if starter_en == "Tailow": starter_en = "Taillow"
        try:
            response = requests.get("https://cobblemon.tools/pokedex/pokemon/"+starter_en.lower()+'/sprite.png')
            cobblemon_icon = Image.open(BytesIO(response.content)).resize((64, 64), Image.Resampling.NEAREST)
            _, _, w, _ = draw.textbbox((0, 0), "Starter : "+starter_fr, font=font_sentence)
            card.paste(cobblemon_icon, (w + 40, 380), cobblemon_icon)
        except (ValueError, UnidentifiedImageError):
            print("Warning: icon for "+starter_en+" could not be found.")
    # Save the card
    output_path = config['PLAYERCARDS']['ImagePath']+"/"+username+".png"
    card.save(output_path)


# Read config
config = configparser.ConfigParser()
config.read('config.ini', encoding='utf8')
if config['INPUT']['Mode'] not in ['manual', 'local', 'ftp', 'sftp']:
    raise Exception("Invalid input mode: "+config['INPUT']['Mode']+". Check the config.")

# Database initialisation
db_path = os.getenv('DB_PATH', 'scoreboard.db')
conn = None
if config['COBBLEMONLEADERBOARDS']['SQLiteOutput'] == "true":
    try:
        conn = init_database(db_path)
    except sqlite3.Error as e:
        print(f"Erreur fatale lors de l'initialisation de la base de données : {e}")
        exit(1)
else:
    print("SQLite désactivé dans la configuration")

# Connect to FTP if activated
ftp_server = None
if config['INPUT']['Mode'] == "ftp":
    ftp_server = ftplib.FTP(config['INPUT']['Host'], open("username.txt", "r").read(), open("password.txt", "r").read())
    ftp_server.encoding = "utf-8"
if config['INPUT']['Mode'] == "sftp":
    transport = paramiko.Transport((config['INPUT']['Host'], int(config['INPUT']['Port'])))
    transport.connect(username=open("username.txt", "r").read().strip(), password=open("password.txt", "r").read().strip())
    ftp_server = paramiko.SFTPClient.from_transport(transport)
    
# Load the vanilla data
print("LOADING VANILLA DATA")
vanilla_df, money_df, waystones_df, advancements_df = loadVanillaData(config['VANILLALEADERBOARD']['CreateCSV'], config['VANILLALEADERBOARD']['CSVPath'], config['INPUT']['Mode'], ftp_server, config['INPUT']['FTPPath'], config['INPUT']['LocalPath'], config['VANILLALEADERBOARD']['CreateCSVMoney'], config['VANILLALEADERBOARD']['CSVPathMoney'])

# Load the Cobblemon data
if config['INPUT']['ImportCobblemon'] == "true":
    print("LOADING COBBLEMON DATA")
    cobblemon_df, cobblemon_df2, cobblemon_df3, cobblemon_df4, cobblemon_df5 = loadCobblemonData(config['GLOBALMATRIX']['CreateCSV'], config['GLOBALMATRIX']['CSVPath'], config['INPUT']['Mode'], ftp_server, config['INPUT']['FTPPath'], config['INPUT']['LocalPath'])


# First leaderboard testing
if config['VANILLALEADERBOARD']['Enable'] == "true":
    getVanillaLeaderboard(vanilla_df, config['VANILLALEADERBOARD']['Category'], config['VANILLALEADERBOARD']['Subcategory'])

# First bestandworst testing
if config['BESTANDWORST']['Enable'] == "true":
    getVanillaBestAndWorst(vanilla_df, config['BESTANDWORST']['Username'], config['BESTANDWORST']['Cleaning'], config['BESTANDWORST']['CleaningValue'])

# Prepare the counting DF
if config['INPUT']['ImportCobblemon'] == "true":
    count_df = cobblemon_df.drop(['caughtTimestamp', 'discoveredTimestamp', 'isShiny'], level=2)
    pokemons_db = pd.read_csv('staticdata/Pokemon.csv')
    legendary_list = pokemons_db.loc[pokemons_db['Legendary'] == True]
    # Don't count twice muk alola and grimer alola
    try:
        if ('muk', 'alola', 'status') in count_df.index and ('mukalolan', 'normal', 'status') in count_df.index:
            new_values = ["CAUGHT" if value1 == "CAUGHT" or value2 == "CAUGHT" else 0 for (value1, value2) in zip(count_df.loc[[('muk', 'alola', 'status')]].values[0], count_df.loc[[('mukalolan', 'normal', 'status')]].values[0])]
            count_df.loc[[('muk', 'alola', 'status')]] = new_values
            count_df.drop('mukalolan', level=0, inplace=True)
    except KeyError:
        print("Warning: Muk Alola form not found in data")

    try:
        if ('grimer', 'alola', 'status') in count_df.index and ('grimeralolan', 'normal', 'status') in count_df.index:
            new_values = ["CAUGHT" if value1 == "CAUGHT" or value2 == "CAUGHT" else 0 for (value1, value2) in zip(count_df.loc[[('grimer', 'alola', 'status')]].values[0], count_df.loc[[('grimeralolan', 'normal', 'status')]].values[0])]
            count_df.loc[[('grimer', 'alola', 'status')]] = new_values
            count_df.drop('grimeralolan', level=0, inplace=True)
    except KeyError:
        print("Warning: Grimer Alola form not found in data")

    # Other counting features
    count_df['times_caught'] = count_df.apply(lambda row: (row == "CAUGHT").sum(), axis=1)
    #print(count_df['times_caught'].sort_values().to_string())
    print("Seen or caught:", len(count_df))
    # Get yet-uncaught pokemons
    caught_count_df = count_df.loc[count_df['times_caught'] > 0]
    print("Caught only:", len(caught_count_df))
    caught_list = (caught_count_df.index.get_level_values(0) + "_" + caught_count_df.index.get_level_values(1)).to_list()
    count_df.drop('times_caught', axis=1, inplace=True)
    uncaught_list = []
    for _, row in pokemons_db.iterrows():
        value = row['Cobblemon'] + "_" + row['Cobblemonform']
        if value not in caught_list:
            uncaught_list.append(value)
    print("Not caught yet (or uncatchable):", len(uncaught_list))
    print(uncaught_list)
    uncaught_excluded_list = list(filter(lambda x: "UNKNOWN" not in x, uncaught_list))
    uncaught_excluded_list.sort()
    print("Not caught yet (or uncatchable), excluding UNKNOWN forms:", len(uncaught_excluded_list))
    print(uncaught_excluded_list)
    # Any pokemons found that are not in pokemon.csv?
    unknown_list = []
    for pokemon in caught_list:
        values = pokemons_db['Cobblemon'] + "_" + pokemons_db['Cobblemonform']
        if pokemon not in values.tolist():
            unknown_list.append(pokemon)
    print("Caught pokemons not found in the db:", len(unknown_list))
    print(unknown_list)

    leaderboards = {}

    # Total leaderboard feature
    player_sum = pd.DataFrame((count_df == "CAUGHT").sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['COBBLEMONLEADERBOARDS']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    #print(player_sum)
    leaderboards["cobblemon_total"] = player_sum
    if config['COBBLEMONLEADERBOARDS']['TotalEnable'] == "true":
        most_pokemons_leaderboard(player_sum, config, "standard", conn)

    # Shiny leaderboard feature
    player_sum = pd.DataFrame(((cobblemon_df == "True") | (cobblemon_df == True)).sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['COBBLEMONLEADERBOARDS']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    #print(player_sum)
    leaderboards["cobblemon_shiny"] = player_sum
    if config['COBBLEMONLEADERBOARDS']['ShinyEnable'] == "true":    
        most_pokemons_leaderboard(player_sum, config, "shiny", conn)

    # Legendary leaderboard feature
    legs = legendary_list['Cobblemon'].tolist()
    leg_count_df = count_df.loc[count_df.index.get_level_values(0).isin(legs)]
    with warnings.catch_warnings():
        warnings.simplefilter(action='ignore', category=FutureWarning)
        leg_count_df = leg_count_df.groupby(level=0).agg(lambda x: "CAUGHT" if "CAUGHT" in x.values else 0)
    #leg_count_df.to_csv("temp.csv")
    player_sum = pd.DataFrame((leg_count_df == "CAUGHT").sum().sort_values())
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['COBBLEMONLEADERBOARDS']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    #print(player_sum)
    leaderboards["cobblemon_legendary"] = player_sum
    if config['COBBLEMONLEADERBOARDS']['LegEnable'] == "true":
        most_pokemons_leaderboard(player_sum, config, "legendary", conn)

    # Money feature
    player_sum = money_df.sort_values('money')
    player_sum['index'] = range(len(player_sum), 0, -1)
    player_sum = player_sum.iloc[::-1]
    ignore_names = [name.strip() for name in config['COBBLEMONLEADERBOARDS']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    #print(player_sum)
    leaderboards["cobblemon_money"] = player_sum
    if config['COBBLEMONLEADERBOARDS']['MoneyEnable'] == "true":
        most_pokemons_leaderboard(player_sum, config, "money",  conn)

if config['TOPIMAGE']['Enable'] == "true":
    leaderboards_to_show = []
    special_list = []
    for leaderboard_type in config['TOPIMAGE']['Leaderboards'].split(','):
        leaderboard_type = leaderboard_type.strip()
        print("Preparing top leaderboard:", leaderboard_type)
        if leaderboard_type.split('/')[0] == "vanilla":
            if leaderboard_type.split('/')[1] == "minecraft:custom":
                if leaderboard_type.split('/')[2] == "minecraft:play_time":
                    leaderboards_to_show.append(getVanillaLeaderboard(vanilla_df, leaderboard_type.split('/')[1], leaderboard_type.split('/')[2], False))
                else:
                    leaderboards_to_show.append(getVanillaLeaderboard(vanilla_df, leaderboard_type.split('/')[1], leaderboard_type.split('/')[2], False))
            elif leaderboard_type.split('/')[1] == "advancements":
                leaderboards_to_show.append(getAdvancementsLeaderboard(advancements_df))
            else:
                leaderboards_to_show.append(getVanillaLeaderboard(vanilla_df, leaderboard_type.split('/')[1], leaderboard_type.split('/')[2], False))
        elif leaderboard_type.split('/')[0] == "cobblemon":
            if config['INPUT']['ImportCobblemon'] != "true":
                print("ERROR: trying to use Cobblemon data in tops feature without importing Cobblemon data. Try setting ImportCobblemon to true.")
                exit()
            if leaderboard_type.split('/')[1] == "pvp":
                leaderboards_to_show.append(getStandardLeaderboard(cobblemon_df2.loc["totalPvPBattleVictoryCount"]))
            elif leaderboard_type.split('/')[1] == "pvw":
                leaderboards_to_show.append(getStandardLeaderboard(cobblemon_df2.loc["totalPvWBattleVictoryCount"]))
            elif leaderboard_type.split('/')[1] == "total":
                leaderboards_to_show.append(leaderboards["cobblemon_total"])
            elif leaderboard_type.split('/')[1] == "shiny":
                leaderboards_to_show.append(leaderboards["cobblemon_shiny"])
            elif leaderboard_type.split('/')[1] == "legendary":
                leaderboards_to_show.append(leaderboards["cobblemon_legendary"])
            elif leaderboard_type.split('/')[1] == "money":
                leaderboards_to_show.append(leaderboards["cobblemon_money"])
            elif leaderboard_type.split('/')[1] == "singletype":
                leaderboards_to_show.append(getCobblemonCaptureCountLeaderboard(cobblemon_df3))
            elif leaderboard_type.split('/')[1] == "waystones":
                leaderboards_to_show.append(getStandardLeaderboard(waystones_df['waystones']))
        # Some leaderboards require a special layout on the image
        if leaderboard_type.split('/')[0] == "cobblemon" and leaderboard_type.split('/')[1] == "singletype":
            if config['INPUT']['ImportCobblemon'] != "true":
                print("ERROR: trying to use Cobblemon data in tops feature without importing Cobblemon data. Try setting ImportCobblemon to true.")
                exit()
            special_list.append("singletype")
        else:
            special_list.append(None)
    # Minecraft-style top feature
    top_image(leaderboards_to_show, config, config['TOPIMAGE']['Titles'].split(","), special_list)

# Duels network graph
if config['PVPNETWORK']['Enable'] == "true":
    PvP_network(cobblemon_df4, config)

# Cobblemon types barchart
if config['TYPESBARCHART']['Enable'] == "true":
    cobblemon_types_barchart(cobblemon_df5, config)

# Cobblemon types barchart
if config['STATSPOKEBALLS']['Enable'] == "true":
    stats_pokeballs(config, ftp_server)

# Custom player cards
if config['PLAYERCARDS']['Enable'] == "true":
    print("Now creating player cards...")
    i = 0
    try:
        df = vanilla_df.drop('zero_count', axis=1)
    except KeyError:
        df = vanilla_df
    for player in df.columns:
        if i%10==0: print("Processing player no", i)
        # Start by downloading the skin image if it is not already downloaded
        destination_path = "images/skins/"+player+".png"
        if not os.path.exists(destination_path):
            os.makedirs(os.path.dirname(destination_path), exist_ok=True)
            response = requests.get("https://minotar.net/armor/bust/"+player+"/500.png")
            if response.status_code == 200:
                with open(destination_path, 'wb') as file:
                    file.write(response.content)
            else:
                print(f"Failed to download skin image. Status code: {response.status_code}")
        try:
            #["Temps de jeu", "Cobblemons attrapés", "Crates lootées", "Duels PvP gagnés", "Advancements"]
            stats_max = [vanilla_df.loc['stats'].loc['minecraft:custom'].loc['minecraft:play_time'].max(),
                        leaderboards["cobblemon_total"].max(),
                        vanilla_df.loc['stats'].loc['minecraft:custom'].loc['lootr:looted_stat'].max(),
                        getStandardLeaderboard(cobblemon_df2.loc["totalPvPBattleVictoryCount"]).max(),
                        getAdvancementsLeaderboard(advancements_df).max()]
            # Scale 0-1
            stats_values = [vanilla_df.loc['stats'].loc['minecraft:custom'].loc['minecraft:play_time'][player]/stats_max[0],
                            (leaderboards["cobblemon_total"].loc[player]/stats_max[1]).iloc[0],
                            vanilla_df.loc['stats'].loc['minecraft:custom'].loc['lootr:looted_stat'][player]/stats_max[2],
                            (getStandardLeaderboard(cobblemon_df2.loc["totalPvPBattleVictoryCount"]).loc[player]/stats_max[3]).iloc[0],
                            (getAdvancementsLeaderboard(advancements_df).loc[player]/stats_max[4]).iloc[0]]
        except KeyError:
            print("Skipped player", player)
            continue
        player_card(config, player, stats_values)
        i += 1

# SQLite close connection
if config['COBBLEMONLEADERBOARDS']['SQLiteOutput'] == "true":
    try:
        conn.close()
        print("Connexion à la base de données SQLite fermée avec succès")
    except sqlite3.Error as e:
        print(f"Erreur lors de la fermeture de la connexion à la base de données : {e}")

# Close the Connection to the FTP/SFTP server
if config['INPUT']['Mode'] == "ftp":
    ftp_server.quit()
if config['INPUT']['Mode'] == "sftp":
    ftp_server.close()

print("Done!")

'''
cobblemon_df2.to_csv('temp/cobblemon2.csv')
cobblemon_df3.to_csv('temp/cobblemon3.csv')
cobblemon_df4.to_csv('temp/cobblemon4.csv')
cobblemon_df5.to_csv('temp/cobblemon5.csv')
money_df.to_csv('temp/money.csv')
waystones_df.to_csv('temp/waystones.csv')
advancements_df.to_csv('temp/advancements.csv')
'''